from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

products = [
    {'id': 'cmmhxuvxh0010pu6jv1m4hmha', 'name': 'Termistor NTC 100K 1% 3950', 'category': 'Componentes Universais', 'image': '/images/products/components-placeholder.svg', 'price_original': None, 'price_promo': None, 'stock': 100},
    {'id': 'cmmhxuvam000wpu6jck11d622', 'name': 'Kit Hotend Completo para Creality CR-10', 'category': 'Componentes Creality', 'image': '/uploads/products/kit-hotend-creality-cr-10.jpg', 'price_original': None, 'price_promo': None, 'stock': 50},
    {'id': 'cmmhxuun2000spu6jrbun5e9g', 'name': 'Kit Aquecedor Cerâmico 60W 360° e Termistor', 'category': 'Componentes Universais', 'image': '/uploads/products/kit-aquecedor-ceramico-60w.jpg', 'price_original': None, 'price_promo': None, 'stock': 100},
    {'id': 'cmmhxuu0c000opu6jig48rtbi', 'name': 'Mesa PEI Texturizada Dupla Face', 'category': 'Componentes Bambu Lab', 'image': '/uploads/products/mesa-pei-texturizada-bambu-lab-h2d.jpg', 'price_original': None, 'price_promo': None, 'stock': 10},
    {'id': 'cmmhxutdm000kpu6j7xjiufx1', 'name': 'Kit Termistor para Bambu Lab A1', 'category': 'Componentes Bambu Lab', 'image': '/images/products/components-placeholder.svg', 'price_original': None, 'price_promo': None, 'stock': 10},
    {'id': 'cmmhxusqx000gpu6jsh5y5kwu', 'name': 'Limpador de Bico Bambu Lab A1', 'category': 'Componentes Bambu Lab', 'image': '/uploads/products/limpador-bico-bambu-lab-a1.jpg', 'price_original': None, 'price_promo': None, 'stock': 50},
    {'id': 'cmmhxus45000cpu6jygv42n58', 'name': 'Bico Nozzle de Aço Endurecido', 'category': 'Componentes Bambu Lab', 'image': '/uploads/products/bico-nozzle-aco-endurecido-bambu-lab-a1.jpg', 'price_original': None, 'price_promo': None, 'stock': 50},
    {'id': 'cmmhxurho0008pu6jdi4ns9de', 'name': 'Capa de Silicone para Bambu Lab A1', 'category': 'Componentes Bambu Lab', 'image': '/uploads/products/capa-silicone-bambu-lab-a1.jpg', 'price_original': None, 'price_promo': None, 'stock': 50},
    {'id': 'cmmhxuqia0004pu6jz9pic8tn', 'name': 'Kit Hotend Completo para Bambu Lab A1', 'category': 'Componentes Bambu Lab', 'image': '/images/products/components-placeholder.svg', 'price_original': None, 'price_promo': None, 'stock': 50}
]

wb = Workbook()
ws = wb.active
ws.title = 'Produtos'

header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, size=14, color='1F4E78')
warning_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
warning_font = Font(color='9C0006')
success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
success_font = Font(color='006100')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Título
ws['A1'] = 'RELATÓRIO DE PRODUTOS IP3D - 15/03/2026'
ws['A1'].font = title_font
ws.merge_cells('A1:G1')
ws.row_dimensions[1].height = 25

# Dados de sumário
row = 3
ws['A' + str(row)] = 'Total de Produtos'
ws['B' + str(row)] = len(products)
row += 1
ws['A' + str(row)] = 'Produtos COM Foto'
ws['B' + str(row)] = sum(1 for p in products if p['image'])
ws['B' + str(row)].font = success_font
row += 1
ws['A' + str(row)] = 'Produtos SEM Foto'
sem_foto = sum(1 for p in products if not p['image'])
ws['B' + str(row)] = sem_foto
if sem_foto > 0:
    ws['B' + str(row)].font = warning_font
row += 1
ws['A' + str(row)] = 'Produtos com Placeholder'
placeholder = sum(1 for p in products if 'placeholder' in p['image'])
ws['B' + str(row)] = placeholder
if placeholder > 0:
    ws['B' + str(row)].font = warning_font
row += 1
ws['A' + str(row)] = 'Produtos SEM Preço'
sem_preco = sum(1 for p in products if not p['price_original'] and not p['price_promo'])
ws['B' + str(row)] = sem_preco
ws['B' + str(row)].font = warning_font

row = 10
headers = ['Nome', 'Categoria', 'Foto Status', 'Imagem Path', 'Preço Original', 'Preço Promo', 'Estoque']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

for idx, product in enumerate(products, 1):
    row += 1
    ws.cell(row=row, column=1).value = product['name']
    ws.cell(row=row, column=2).value = product['category']

    foto = ws.cell(row=row, column=3)
    if 'placeholder' in product['image']:
        foto.value = 'Placeholder'
        foto.fill = warning_fill
        foto.font = warning_font
    else:
        foto.value = 'Custom'
        foto.fill = success_fill
        foto.font = success_font

    ws.cell(row=row, column=4).value = product['image']
    ws.cell(row=row, column=5).value = 'SEM PRECO' if not product['price_original'] else product['price_original']
    ws.cell(row=row, column=6).value = 'SEM PRECO' if not product['price_promo'] else product['price_promo']
    ws.cell(row=row, column=7).value = product['stock']

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 10

wb.save('D:\\IP3D Node\\PRODUTOS-ANALISE-COMPLETA.xlsx')
print('✅ Arquivo criado: PRODUTOS-ANALISE-COMPLETA.xlsx')
print('')
print('RESUMO:')
print('   Total de produtos: 9')
print('   Produtos COM foto: 9 (100%)')
print('   Produtos SEM foto: 0')
print('   Produtos com placeholder: 3')
print('   Produtos SEM preco: 9 (PROBLEMA CRITICO)')
